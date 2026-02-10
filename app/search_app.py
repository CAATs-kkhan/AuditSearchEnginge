"""
Audit Document Search Engine - Phase 9 with AI Features
========================================================
A search tool for auditors with NER, Topic Modelling, and Advanced AI Features.

Features:
- Simple, Boolean, Wildcard, Fuzzy search
- Named Entity Recognition (NER)
- PDF Table Extraction
- Topic Modelling (LDA)
- Sentiment & Risk Analysis
- AI Features (NEW!):
  - Document Summarization (using BART)
  - Q&A Chat with Documents (RAG with ChromaDB)
  - Anomaly Detection (Isolation Forest)
- Export to Excel

How to run:
    streamlit run search_app.py

New Dependencies for AI Features:
    pip install transformers torch sentence-transformers chromadb
"""

import streamlit as st
import os
import json
import re
import io
from pathlib import Path
from datetime import datetime
from collections import Counter

# Document processing imports
try:
    from pypdf import PdfReader
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    from docx import Document as DocxDocument
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False

try:
    import openpyxl
    from openpyxl import Workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Try to import spaCy for advanced NER
try:
    import spacy
    SPACY_AVAILABLE = True
except ImportError:
    SPACY_AVAILABLE = False

# Try to import pdfplumber for table extraction
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

# Try to import scikit-learn for topic modelling
try:
    from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
    from sklearn.decomposition import LatentDirichletAllocation
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False

# Try to import TextBlob for sentiment analysis
try:
    from textblob import TextBlob
    TEXTBLOB_AVAILABLE = True
except ImportError:
    TEXTBLOB_AVAILABLE = False

# Try to import pytesseract for OCR
try:
    import pytesseract
    from PIL import Image
    # Set Tesseract path (adjust if different)
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False

# Try to import pdf2image for PDF to image conversion
try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False

# Try to import transformers for summarization
try:
    from transformers import AutoTokenizer, AutoModelForSeq2SeqLM
    TRANSFORMERS_AVAILABLE = True
except ImportError:
    TRANSFORMERS_AVAILABLE = False

# Try to import sentence-transformers for embeddings
try:
    from sentence_transformers import SentenceTransformer
    SENTENCE_TRANSFORMERS_AVAILABLE = True
except ImportError:
    SENTENCE_TRANSFORMERS_AVAILABLE = False

# Try to import faiss for vector storage (replacing chromadb due to Python 3.14 issues)
FAISS_AVAILABLE = False
try:
    import faiss
    FAISS_AVAILABLE = True
except ImportError:
    pass

# Legacy chromadb flag for compatibility
CHROMADB_AVAILABLE = FAISS_AVAILABLE
CHROMADB_ERROR = None if FAISS_AVAILABLE else "Using FAISS instead of ChromaDB"

# Try to import numpy for anomaly detection
try:
    import numpy as np
    NUMPY_AVAILABLE = True
except ImportError:
    NUMPY_AVAILABLE = False

# Find Poppler path for pdf2image (Windows)
POPPLER_PATH = None
if PDF2IMAGE_AVAILABLE:
    import glob
    # Common Poppler installation paths on Windows
    poppler_patterns = [
        r'C:\poppler\poppler-*\Library\bin',
        r'C:\poppler-*\Library\bin',
        r'C:\poppler\poppler-*\bin',
        r'C:\poppler-*\bin',
        r'C:\poppler\Library\bin',
        r'C:\poppler\bin',
        r'C:\Program Files\poppler*\bin',
        r'C:\Program Files\poppler*\Library\bin',
    ]
    for pattern in poppler_patterns:
        matches = glob.glob(pattern)
        if matches:
            POPPLER_PATH = matches[0]
            break

# =============================================================================
# CONFIGURATION
# =============================================================================

BASE_DIR = Path(__file__).parent.parent
DOCUMENTS_FOLDER = BASE_DIR / "documents"
INDEX_FILE = BASE_DIR / "search_index.json"
ENTITIES_FILE = BASE_DIR / "entities_index.json"
TOPICS_FILE = BASE_DIR / "topics_index.json"
SENTIMENT_FILE = BASE_DIR / "sentiment_index.json"
SUMMARIES_FILE = BASE_DIR / "summaries_index.json"
CHROMA_DIR = BASE_DIR / "chroma_db"
DOCUMENTS_FOLDER.mkdir(exist_ok=True)

# =============================================================================
# DOCUMENT TEXT EXTRACTION
# =============================================================================

def extract_text_from_scanned_pdf(file_path):
    """Extract text from scanned PDF using Tesseract OCR."""
    if not TESSERACT_AVAILABLE or not PDF2IMAGE_AVAILABLE:
        return ""

    try:
        # Convert PDF pages to images (use poppler_path on Windows)
        if POPPLER_PATH:
            images = convert_from_path(file_path, dpi=200, poppler_path=POPPLER_PATH)
        else:
            images = convert_from_path(file_path, dpi=200)

        all_text = []
        for i, image in enumerate(images):
            # Run OCR on each page
            page_text = pytesseract.image_to_string(image)
            all_text.append(page_text)

        return '\n\n'.join(all_text)
    except Exception as e:
        return f"[OCR Error: {str(e)}]"


def extract_text_from_pdf(file_path):
    """Extract text from a PDF file. Falls back to OCR for scanned PDFs."""
    if not PDF_SUPPORT:
        return "[PDF support not installed]"

    text = ""
    try:
        reader = PdfReader(file_path)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

        # If no text extracted, try OCR
        if not text.strip() and TESSERACT_AVAILABLE and PDF2IMAGE_AVAILABLE:
            text = extract_text_from_scanned_pdf(file_path)

        return text
    except Exception as e:
        return f"[Error reading PDF: {str(e)}]"


def extract_text_from_docx(file_path):
    """Extract text from a Word document."""
    if not DOCX_SUPPORT:
        return "[Word support not installed]"
    text = ""
    try:
        doc = DocxDocument(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        text = f"[Error reading Word file: {e}]"
    return text


def extract_text_from_excel(file_path):
    """Extract text from an Excel file."""
    if not EXCEL_SUPPORT:
        return "[Excel support not installed]"
    text = ""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet in wb.worksheets:
            text += f"\n--- Sheet: {sheet.title} ---\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
    except Exception as e:
        text = f"[Error reading Excel file: {e}]"
    return text


def extract_text_from_txt(file_path):
    """Extract text from a plain text file."""
    encodings = ['utf-8', 'utf-16', 'latin-1', 'cp1252']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                return f.read()
        except:
            continue
    return "[Could not read text file]"


def extract_text(file_path):
    """Extract text from a document based on its type."""
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    if suffix == '.pdf':
        return extract_text_from_pdf(file_path)
    elif suffix in ['.docx', '.doc']:
        return extract_text_from_docx(file_path)
    elif suffix in ['.xlsx', '.xls']:
        return extract_text_from_excel(file_path)
    elif suffix in ['.txt', '.text', '.md']:
        return extract_text_from_txt(file_path)
    else:
        return f"[Unsupported file type: {suffix}]"


# =============================================================================
# NAMED ENTITY RECOGNITION (NER) - Pattern Based
# =============================================================================

def extract_entities_pattern(text):
    """
    Extract entities using regex patterns.
    Works without any external NLP library.
    """
    entities = {
        "persons": [],
        "organizations": [],
        "locations": [],
        "dates": [],
        "money": [],
        "emails": [],
        "phones": [],
        "percentages": []
    }

    # --- EMAIL ADDRESSES ---
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    entities["emails"] = list(set(re.findall(email_pattern, text)))

    # --- PHONE NUMBERS ---
    phone_patterns = [
        r'\b\d{3}[-.\s]?\d{3}[-.\s]?\d{4}\b',  # 123-456-7890
        r'\b\(\d{3}\)\s*\d{3}[-.\s]?\d{4}\b',   # (123) 456-7890
        r'\b\+\d{1,3}[-.\s]?\d{3}[-.\s]?\d{3}[-.\s]?\d{4}\b',  # +1-123-456-7890
        r'\b\d{4}[-.\s]?\d{7}\b',  # 0321-1234567 (Pakistani format)
    ]
    phones = []
    for pattern in phone_patterns:
        phones.extend(re.findall(pattern, text))
    entities["phones"] = list(set(phones))

    # --- DATES ---
    date_patterns = [
        # DD/MM/YYYY or DD-MM-YYYY
        r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b',
        # Month DD, YYYY
        r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4}\b',
        # DD Month YYYY
        r'\b\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}\b',
        # Month YYYY
        r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}\b',
        # Abbreviated months
        r'\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[.]?\s+\d{1,2},?\s+\d{4}\b',
        r'\b\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[.]?\s+\d{4}\b',
    ]
    dates = []
    for pattern in date_patterns:
        dates.extend(re.findall(pattern, text, re.IGNORECASE))
    entities["dates"] = list(set(dates))

    # --- MONEY ---
    money_patterns = [
        # Dollar amounts
        r'\$\s?\d{1,3}(?:,\d{3})*(?:\.\d{2})?(?:\s?(?:million|billion|thousand|M|B|K))?',
        # USD amounts
        r'USD\s?\d{1,3}(?:,\d{3})*(?:\.\d{2})?(?:\s?(?:million|billion|thousand|M|B|K))?',
        # Pakistani Rupees
        r'(?:Rs\.?|PKR)\s?\d{1,3}(?:,\d{3})*(?:\.\d{2})?(?:\s?(?:million|billion|thousand|lac|lakh|crore))?',
        # Euro
        r'€\s?\d{1,3}(?:,\d{3})*(?:\.\d{2})?(?:\s?(?:million|billion|thousand|M|B|K))?',
        # British Pound
        r'£\s?\d{1,3}(?:,\d{3})*(?:\.\d{2})?(?:\s?(?:million|billion|thousand|M|B|K))?',
        # Generic with currency words
        r'\d{1,3}(?:,\d{3})*(?:\.\d{2})?\s?(?:dollars|rupees|euros|pounds)',
    ]
    money = []
    for pattern in money_patterns:
        money.extend(re.findall(pattern, text, re.IGNORECASE))
    entities["money"] = list(set(money))

    # --- PERCENTAGES ---
    percentage_pattern = r'\b\d+(?:\.\d+)?%'
    entities["percentages"] = list(set(re.findall(percentage_pattern, text)))

    # --- PERSONS (Name patterns) ---
    # Look for common name patterns (Title + Name)
    person_patterns = [
        # Mr./Mrs./Ms./Dr. + Name
        r'\b(?:Mr\.|Mrs\.|Ms\.|Dr\.|Prof\.|Sir|Madam)\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*',
        # Names with common Pakistani/Arabic honorifics
        r'\b(?:Muhammad|Mohammad|Ahmed|Ali|Khan|Sheikh|Malik|Chaudhry|Mirza)\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*',
    ]
    persons = []
    for pattern in person_patterns:
        persons.extend(re.findall(pattern, text))

    # Also find capitalized word sequences (potential names)
    # Two or more capitalized words together
    name_pattern = r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3}\b'
    potential_names = re.findall(name_pattern, text)

    # Filter out common non-name phrases
    non_names = {'The', 'This', 'That', 'These', 'Those', 'What', 'When', 'Where',
                 'Which', 'While', 'During', 'After', 'Before', 'January', 'February',
                 'March', 'April', 'May', 'June', 'July', 'August', 'September',
                 'October', 'November', 'December', 'Monday', 'Tuesday', 'Wednesday',
                 'Thursday', 'Friday', 'Saturday', 'Sunday', 'New York', 'United States',
                 'United Kingdom', 'Supreme Court', 'High Court', 'National Assembly'}

    for name in potential_names:
        first_word = name.split()[0]
        if first_word not in non_names and len(name) > 5:
            persons.append(name)

    persons.extend([p for p in persons if p])
    entities["persons"] = list(set(persons))[:50]  # Limit to top 50

    # --- ORGANIZATIONS ---
    org_patterns = [
        # Companies with common suffixes
        r'\b[A-Z][A-Za-z]+(?:\s+[A-Z][A-Za-z]+)*\s+(?:Ltd|Limited|Inc|Corporation|Corp|Company|Co|LLC|LLP|PLC|Pvt|Private)\b',
        # Government bodies
        r'\b(?:Ministry|Department|Bureau|Agency|Authority|Commission|Committee|Board|Council|Office)\s+(?:of\s+)?[A-Z][A-Za-z]+(?:\s+[A-Z]?[A-Za-z]+)*',
        # Banks
        r'\b[A-Z][A-Za-z]+(?:\s+[A-Z][A-Za-z]+)*\s+Bank\b',
        # Universities/Institutions
        r'\b(?:University|Institute|College|School|Academy)\s+(?:of\s+)?[A-Z][A-Za-z]+(?:\s+[A-Z]?[A-Za-z]+)*',
        r'\b[A-Z][A-Za-z]+(?:\s+[A-Z][A-Za-z]+)*\s+(?:University|Institute|College)\b',
    ]
    organizations = []
    for pattern in org_patterns:
        organizations.extend(re.findall(pattern, text))
    entities["organizations"] = list(set(organizations))[:50]

    # --- LOCATIONS ---
    # Common Pakistani cities and locations (English)
    pakistan_locations = [
        'Islamabad', 'Karachi', 'Lahore', 'Peshawar', 'Quetta', 'Faisalabad',
        'Rawalpindi', 'Multan', 'Hyderabad', 'Gujranwala', 'Sialkot', 'Bahawalpur',
        'Sargodha', 'Sukkur', 'Larkana', 'Sheikhupura', 'Jhang', 'Rahim Yar Khan',
        'Gujrat', 'Mardan', 'Kasur', 'Mingora', 'Dera Ghazi Khan', 'Sahiwal',
        'Nawabshah', 'Okara', 'Mirpur Khas', 'Chiniot', 'Kamoke', 'Sadiqabad',
        'Burewala', 'Jacobabad', 'Muzaffargarh', 'Muridke', 'Jhelum', 'Shikarpur',
        'Hafizabad', 'Kohat', 'Khanewal', 'Daska', 'Punjab', 'Sindh', 'KPK',
        'Balochistan', 'Gilgit', 'Baltistan', 'Azad Kashmir', 'Pakistan'
    ]

    # Pakistani cities in Urdu
    pakistan_locations_urdu = [
        'اسلام آباد', 'کراچی', 'لاہور', 'پشاور', 'کوئٹہ', 'فیصل آباد',
        'راولپنڈی', 'ملتان', 'حیدرآباد', 'گوجرانوالہ', 'سیالکوٹ', 'بہاولپور',
        'سرگودھا', 'سکھر', 'لاڑکانہ', 'شیخوپورہ', 'جھنگ', 'رحیم یار خان',
        'گجرات', 'مردان', 'قصور', 'مینگورہ', 'ڈیرہ غازی خان', 'ساہیوال',
        'نوابشاہ', 'اوکاڑہ', 'میرپور خاص', 'چنیوٹ', 'کاموکی', 'صادق آباد',
        'پنجاب', 'سندھ', 'خیبر پختونخوا', 'بلوچستان', 'گلگت', 'بلتستان',
        'آزاد کشمیر', 'پاکستان'
    ]

    # International locations
    international_locations = [
        'United States', 'United Kingdom', 'China', 'India', 'Saudi Arabia',
        'UAE', 'Dubai', 'London', 'New York', 'Washington', 'Beijing',
        'Tokyo', 'Paris', 'Berlin', 'Moscow', 'Sydney', 'Toronto',
        'Canada', 'Australia', 'Germany', 'France', 'Japan', 'Singapore',
        'Malaysia', 'Indonesia', 'Bangladesh', 'Sri Lanka', 'Nepal',
        'Afghanistan', 'Iran', 'Iraq', 'Qatar', 'Kuwait', 'Bahrain', 'Oman'
    ]

    # International locations in Urdu
    international_locations_urdu = [
        'امریکہ', 'برطانیہ', 'چین', 'بھارت', 'سعودی عرب',
        'متحدہ عرب امارات', 'دبئی', 'لندن', 'نیویارک', 'واشنگٹن',
        'ٹوکیو', 'پیرس', 'برلن', 'ماسکو', 'سڈنی', 'ٹورنٹو',
        'کینیڈا', 'آسٹریلیا', 'جرمنی', 'فرانس', 'جاپان', 'سنگاپور',
        'ملائیشیا', 'انڈونیشیا', 'بنگلہ دیش', 'سری لنکا', 'نیپال',
        'افغانستان', 'ایران', 'عراق', 'قطر', 'کویت', 'بحرین', 'عمان'
    ]

    all_known_locations = (pakistan_locations + pakistan_locations_urdu +
                          international_locations + international_locations_urdu)

    # Only add locations that are in our known list (more accurate)
    locations = []
    for loc in all_known_locations:
        if loc in text:
            locations.append(loc)

    # Remove duplicates and limit
    entities["locations"] = list(set(locations))[:50]

    return entities


def extract_entities_spacy(text, nlp):
    """
    Extract entities using spaCy NLP library.
    More accurate but requires spaCy installation.
    """
    entities = {
        "persons": [],
        "organizations": [],
        "locations": [],
        "dates": [],
        "money": [],
        "emails": [],
        "phones": [],
        "percentages": []
    }

    # Process with spaCy (limit text length for performance)
    max_length = 100000
    if len(text) > max_length:
        text = text[:max_length]

    doc = nlp(text)

    for ent in doc.ents:
        if ent.label_ == "PERSON":
            entities["persons"].append(ent.text)
        elif ent.label_ == "ORG":
            entities["organizations"].append(ent.text)
        elif ent.label_ in ("GPE", "LOC"):
            entities["locations"].append(ent.text)
        elif ent.label_ == "DATE":
            entities["dates"].append(ent.text)
        elif ent.label_ == "MONEY":
            entities["money"].append(ent.text)
        elif ent.label_ == "PERCENT":
            entities["percentages"].append(ent.text)

    # Also use pattern-based for emails and phones (spaCy doesn't detect these)
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    entities["emails"] = list(set(re.findall(email_pattern, text)))

    phone_patterns = [
        r'\b\d{3}[-.\s]?\d{3}[-.\s]?\d{4}\b',
        r'\b\(\d{3}\)\s*\d{3}[-.\s]?\d{4}\b',
        r'\b\d{4}[-.\s]?\d{7}\b',
    ]
    phones = []
    for pattern in phone_patterns:
        phones.extend(re.findall(pattern, text))
    entities["phones"] = list(set(phones))

    # Deduplicate
    for key in entities:
        entities[key] = list(set(entities[key]))[:50]

    return entities


# =============================================================================
# SEARCH INDEX MANAGEMENT
# =============================================================================

def load_index():
    """Load the search index from file."""
    if INDEX_FILE.exists():
        with open(INDEX_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"documents": {}}


def save_index(index):
    """Save the search index to file."""
    with open(INDEX_FILE, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=2)


def load_entities():
    """Load entities index from file."""
    if ENTITIES_FILE.exists():
        with open(ENTITIES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"documents": {}, "global": {}}


def save_entities(entities_index):
    """Save entities index to file."""
    with open(ENTITIES_FILE, 'w', encoding='utf-8') as f:
        json.dump(entities_index, f, ensure_ascii=False, indent=2)


def add_document_to_index(file_path, index, entities_index, nlp=None):
    """Add a document to the search index and extract entities."""
    file_path = Path(file_path)
    doc_id = file_path.name
    text = extract_text(file_path)

    # Count words for statistics
    words = re.findall(r'\b\w+\b', text.lower())
    word_count = len(words)

    # Extract entities
    if nlp and SPACY_AVAILABLE:
        entities = extract_entities_spacy(text, nlp)
    else:
        entities = extract_entities_pattern(text)

    index["documents"][doc_id] = {
        "path": str(file_path),
        "name": file_path.name,
        "text": text,
        "word_count": word_count,
        "indexed_at": datetime.now().isoformat(),
        "size": file_path.stat().st_size if file_path.exists() else 0
    }

    entities_index["documents"][doc_id] = entities

    return text, entities


def index_all_documents(use_spacy=False):
    """Index all documents in the documents folder."""
    index = {"documents": {}}
    entities_index = {"documents": {}, "global": {
        "persons": [],
        "organizations": [],
        "locations": [],
        "dates": [],
        "money": []
    }}

    # Load spaCy model if available and requested
    nlp = None
    if use_spacy and SPACY_AVAILABLE:
        try:
            nlp = spacy.load("en_core_web_sm")
        except:
            nlp = None

    supported_extensions = ['.pdf', '.docx', '.doc', '.xlsx', '.xls', '.txt', '.md']

    for file_path in DOCUMENTS_FOLDER.iterdir():
        if file_path.suffix.lower() in supported_extensions:
            _, entities = add_document_to_index(file_path, index, entities_index, nlp)

            # Aggregate global entities
            for key in ["persons", "organizations", "locations", "dates", "money"]:
                entities_index["global"][key].extend(entities.get(key, []))

    # Count frequencies for global entities
    for key in entities_index["global"]:
        counter = Counter(entities_index["global"][key])
        entities_index["global"][key] = counter.most_common(100)

    save_index(index)
    save_entities(entities_index)
    return index, entities_index


# =============================================================================
# ENHANCED SEARCH FUNCTIONS
# =============================================================================

def simple_search(query, index, case_sensitive=False):
    """Simple keyword search across all documents."""
    results = []
    if not case_sensitive:
        query = query.lower()

    for doc_id, doc_data in index.get("documents", {}).items():
        text = doc_data.get("text", "")
        search_text = text if case_sensitive else text.lower()

        if query in search_text:
            pos = search_text.find(query)
            start = max(0, pos - 100)
            end = min(len(text), pos + len(query) + 100)
            snippet = text[start:end]
            count = search_text.count(query)

            results.append({
                "document": doc_data.get("name"),
                "path": doc_data.get("path"),
                "snippet": snippet,
                "matches": count,
                "query_terms": [query]
            })

    results.sort(key=lambda x: x["matches"], reverse=True)
    return results


def fuzzy_search(query, index):
    """Fuzzy search that finds partial matches."""
    results = []
    query_lower = query.lower()
    query_words = query_lower.split()

    for doc_id, doc_data in index.get("documents", {}).items():
        text = doc_data.get("text", "").lower()
        original_text = doc_data.get("text", "")
        matches = 0
        snippets = []
        matched_terms = []

        for word in query_words:
            if len(word) >= 3:
                if word in text:
                    matches += 1
                    matched_terms.append(word)
                    pos = text.find(word)
                    start = max(0, pos - 50)
                    end = min(len(text), pos + len(word) + 50)
                    snippets.append(original_text[start:end])

        if matches > 0:
            results.append({
                "document": doc_data.get("name"),
                "path": doc_data.get("path"),
                "snippet": "...".join(snippets[:2]),
                "matches": matches,
                "relevance": matches / len(query_words) if query_words else 0,
                "query_terms": matched_terms
            })

    results.sort(key=lambda x: x["relevance"], reverse=True)
    return results


def boolean_search(query, index, case_sensitive=False):
    """Boolean search supporting AND, OR, NOT operators."""
    results = []
    query_upper = query.upper()
    has_and = " AND " in query_upper
    has_or = " OR " in query_upper
    has_not = " NOT " in query_upper

    for doc_id, doc_data in index.get("documents", {}).items():
        text = doc_data.get("text", "")
        search_text = text if case_sensitive else text.lower()
        query_for_search = query if case_sensitive else query.lower()

        include_doc = False
        matched_terms = []

        if has_and:
            terms = [t.strip().lower() for t in re.split(r'\s+AND\s+', query, flags=re.IGNORECASE)]
            if all(term in search_text for term in terms):
                include_doc = True
                matched_terms = terms
        elif has_or:
            terms = [t.strip().lower() for t in re.split(r'\s+OR\s+', query, flags=re.IGNORECASE)]
            for term in terms:
                if term in search_text:
                    include_doc = True
                    matched_terms.append(term)
        elif has_not:
            parts = re.split(r'\s+NOT\s+', query, flags=re.IGNORECASE)
            if len(parts) >= 2:
                must_have = parts[0].strip().lower()
                must_not_have = parts[1].strip().lower()
                if must_have in search_text and must_not_have not in search_text:
                    include_doc = True
                    matched_terms = [must_have]
        else:
            if query_for_search in search_text:
                include_doc = True
                matched_terms = [query_for_search]

        if include_doc:
            if matched_terms:
                pos = search_text.find(matched_terms[0])
                start = max(0, pos - 100)
                end = min(len(text), pos + len(matched_terms[0]) + 100)
                snippet = text[start:end]
            else:
                snippet = text[:200]

            total_matches = sum(search_text.count(term) for term in matched_terms)

            results.append({
                "document": doc_data.get("name"),
                "path": doc_data.get("path"),
                "snippet": snippet,
                "matches": total_matches,
                "query_terms": matched_terms
            })

    results.sort(key=lambda x: x["matches"], reverse=True)
    return results


def wildcard_search(query, index, case_sensitive=False):
    """Wildcard search using * as a wildcard."""
    results = []
    pattern = re.escape(query).replace(r'\*', '.*')

    if not case_sensitive:
        regex = re.compile(pattern, re.IGNORECASE)
    else:
        regex = re.compile(pattern)

    for doc_id, doc_data in index.get("documents", {}).items():
        text = doc_data.get("text", "")
        matches = regex.findall(text)

        if matches:
            unique_matches = list(set(matches))
            match = regex.search(text)
            if match:
                pos = match.start()
                start = max(0, pos - 100)
                end = min(len(text), match.end() + 100)
                snippet = text[start:end]
            else:
                snippet = text[:200]

            results.append({
                "document": doc_data.get("name"),
                "path": doc_data.get("path"),
                "snippet": snippet,
                "matches": len(matches),
                "query_terms": unique_matches[:5],
                "all_matches": unique_matches
            })

    results.sort(key=lambda x: x["matches"], reverse=True)
    return results


def entity_search(entity_type, entity_value, index, entities_index):
    """Search for documents containing a specific entity."""
    results = []

    for doc_id, doc_data in index.get("documents", {}).items():
        doc_entities = entities_index.get("documents", {}).get(doc_id, {})
        entities_of_type = doc_entities.get(entity_type, [])

        # Check if entity exists in this document
        entity_lower = entity_value.lower()
        matches = [e for e in entities_of_type if entity_lower in e.lower()]

        if matches:
            text = doc_data.get("text", "")
            # Find snippet around the entity
            search_text = text.lower()
            pos = search_text.find(entity_lower)
            if pos >= 0:
                start = max(0, pos - 100)
                end = min(len(text), pos + len(entity_value) + 100)
                snippet = text[start:end]
            else:
                snippet = text[:200]

            results.append({
                "document": doc_data.get("name"),
                "path": doc_data.get("path"),
                "snippet": snippet,
                "matches": len(matches),
                "query_terms": matches
            })

    results.sort(key=lambda x: x["matches"], reverse=True)
    return results


# =============================================================================
# HIGHLIGHT AND EXPORT FUNCTIONS
# =============================================================================

def highlight_text(text, terms, highlight_color="#FFFF00"):
    """Highlight search terms in text using HTML."""
    highlighted = text
    for term in terms:
        if term:
            pattern = re.compile(re.escape(term), re.IGNORECASE)
            highlighted = pattern.sub(
                f'<mark style="background-color: {highlight_color}; padding: 2px;">{term}</mark>',
                highlighted
            )
    return highlighted


def export_results_to_excel(results, query):
    """Export search results to Excel file."""
    if not EXCEL_SUPPORT:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"

    ws['A1'] = "Search Query:"
    ws['B1'] = query
    ws['A2'] = "Export Date:"
    ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A3'] = "Total Results:"
    ws['B3'] = len(results)

    headers = ["#", "Document", "Matches", "Matched Terms", "Snippet", "File Path"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=header)

    for i, result in enumerate(results, 1):
        ws.cell(row=5+i, column=1, value=i)
        ws.cell(row=5+i, column=2, value=result.get("document", ""))
        ws.cell(row=5+i, column=3, value=result.get("matches", 0))
        ws.cell(row=5+i, column=4, value=", ".join(result.get("query_terms", [])))
        ws.cell(row=5+i, column=5, value=result.get("snippet", "")[:500])
        ws.cell(row=5+i, column=6, value=result.get("path", ""))

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 80
    ws.column_dimensions['F'].width = 50

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_entities_to_excel(entities_index):
    """Export entities summary to Excel file."""
    if not EXCEL_SUPPORT:
        return None

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary['A1'] = "Entity Extraction Summary"
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    row = 4
    for entity_type, entities in entities_index.get("global", {}).items():
        ws_summary.cell(row=row, column=1, value=entity_type.upper())
        ws_summary.cell(row=row, column=2, value=f"{len(entities)} unique")
        row += 1

    # Create sheet for each entity type
    for entity_type in ["persons", "organizations", "locations", "dates", "money"]:
        ws = wb.create_sheet(title=entity_type.capitalize())
        ws['A1'] = "Entity"
        ws['B1'] = "Frequency"

        entities = entities_index.get("global", {}).get(entity_type, [])
        for i, (entity, count) in enumerate(entities[:100], 2):
            ws.cell(row=i, column=1, value=entity)
            ws.cell(row=i, column=2, value=count)

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_search_statistics(results, index):
    """Calculate search statistics."""
    total_docs = len(index.get("documents", {}))
    docs_found = len(results)
    total_matches = sum(r.get("matches", 0) for r in results)
    all_terms = []
    for r in results:
        all_terms.extend(r.get("query_terms", []))
    term_freq = Counter(all_terms)

    return {
        "total_documents": total_docs,
        "documents_found": docs_found,
        "total_matches": total_matches,
        "match_percentage": (docs_found / total_docs * 100) if total_docs > 0 else 0,
        "term_frequency": term_freq.most_common(10)
    }


# =============================================================================
# PDF TABLE EXTRACTION FUNCTIONS
# =============================================================================

def extract_tables_from_pdf(pdf_path):
    """
    Extract all tables from a PDF file using pdfplumber.
    Returns a list of tables, each table is a list of rows.
    """
    if not PDFPLUMBER_AVAILABLE:
        return [], "pdfplumber not installed"

    tables = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                page_tables = page.extract_tables()
                for table_num, table in enumerate(page_tables, 1):
                    if table and len(table) > 0:
                        # Clean the table data
                        cleaned_table = []
                        for row in table:
                            if row:
                                cleaned_row = [str(cell).strip() if cell else "" for cell in row]
                                # Only add rows that have some content
                                if any(cell for cell in cleaned_row):
                                    cleaned_table.append(cleaned_row)

                        if cleaned_table:
                            tables.append({
                                "page": page_num,
                                "table_num": table_num,
                                "data": cleaned_table,
                                "rows": len(cleaned_table),
                                "cols": len(cleaned_table[0]) if cleaned_table else 0
                            })
        return tables, None
    except Exception as e:
        return [], str(e)


def extract_tables_from_multiple_pdfs(pdf_files):
    """
    Extract tables from multiple PDF files.
    Returns a dictionary with file names as keys and tables as values.
    """
    all_tables = {}
    errors = []

    for pdf_file in pdf_files:
        file_name = pdf_file.name if hasattr(pdf_file, 'name') else str(pdf_file)
        tables, error = extract_tables_from_pdf(pdf_file)

        if error:
            errors.append(f"{file_name}: {error}")
        else:
            all_tables[file_name] = tables

    return all_tables, errors


def clean_cell_value(value):
    """Clean cell value to be Excel-compatible."""
    if value is None:
        return ""

    # Convert to string
    value = str(value)

    # Remove or replace invalid XML characters
    # Excel uses XML internally, so these characters cause problems
    invalid_chars = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')
    value = invalid_chars.sub('', value)

    # Replace newlines with spaces
    value = value.replace('\n', ' ').replace('\r', ' ')

    # Limit cell length (Excel has a limit of 32767 characters per cell)
    if len(value) > 32000:
        value = value[:32000] + "..."

    return value.strip()


def clean_sheet_name(name):
    """Clean sheet name to be Excel-compatible."""
    # Remove invalid characters for Excel sheet names
    invalid_chars = r'[\\/*?:\[\]]'
    name = re.sub(invalid_chars, '_', name)

    # Remove leading/trailing quotes and spaces
    name = name.strip("'\" ")

    # Excel sheet names max 31 characters
    if len(name) > 31:
        name = name[:28] + "..."

    # Ensure name is not empty
    if not name:
        name = "Sheet"

    return name


def export_tables_to_excel(all_tables, mode="separate_sheets"):
    """
    Export extracted tables to Excel.

    Modes:
    - "separate_sheets": Each PDF gets its own sheet
    - "single_sheet": All tables merged into one sheet
    """
    if not EXCEL_SUPPORT:
        return None

    wb = Workbook()

    if mode == "separate_sheets":
        # Use the default sheet for the first file, then create new sheets
        first_file = True

        for file_name, tables in all_tables.items():
            # Create sheet name
            sheet_name = clean_sheet_name(file_name.replace('.pdf', '').replace('.PDF', ''))

            if first_file:
                ws = wb.active
                ws.title = sheet_name
                first_file = False
            else:
                ws = wb.create_sheet(title=sheet_name)

            current_row = 1
            ws.cell(row=current_row, column=1, value=f"Tables extracted from: {clean_cell_value(file_name)}")
            ws.cell(row=current_row + 1, column=1, value=f"Extracted on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            current_row += 3

            for table_info in tables:
                # Table header
                ws.cell(row=current_row, column=1,
                       value=f"Table {table_info['table_num']} from Page {table_info['page']}")
                current_row += 1

                # Table data
                for row_data in table_info['data']:
                    for col_idx, cell_value in enumerate(row_data, 1):
                        ws.cell(row=current_row, column=col_idx, value=clean_cell_value(cell_value))
                    current_row += 1

                current_row += 2  # Space between tables

    elif mode == "single_sheet":
        # All tables in one sheet
        ws = wb.active
        ws.title = "All Tables"

        current_row = 1
        ws.cell(row=current_row, column=1, value="Consolidated Tables from Multiple PDFs")
        ws.cell(row=current_row + 1, column=1,
               value=f"Extracted on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=current_row + 2, column=1, value=f"Total files: {len(all_tables)}")
        current_row += 4

        for file_name, tables in all_tables.items():
            ws.cell(row=current_row, column=1, value=f"FILE: {clean_cell_value(file_name)}")
            current_row += 1

            for table_info in tables:
                ws.cell(row=current_row, column=1,
                       value=f"Table {table_info['table_num']} (Page {table_info['page']})")
                current_row += 1

                for row_data in table_info['data']:
                    for col_idx, cell_value in enumerate(row_data, 1):
                        ws.cell(row=current_row, column=col_idx, value=clean_cell_value(cell_value))
                    current_row += 1

                current_row += 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# TOPIC MODELLING FUNCTIONS
# =============================================================================

def load_topics():
    """Load topics index from file."""
    if TOPICS_FILE.exists():
        with open(TOPICS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"topics": [], "document_topics": {}}


def save_topics(topics_index):
    """Save topics index to file."""
    with open(TOPICS_FILE, 'w', encoding='utf-8') as f:
        json.dump(topics_index, f, ensure_ascii=False, indent=2)


def preprocess_text_for_topics(text):
    """Preprocess text for topic modelling."""
    # Convert to lowercase
    text = text.lower()

    # Remove special characters and numbers
    text = re.sub(r'[^a-zA-Z\s]', ' ', text)

    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text).strip()

    # Remove short words (less than 3 characters)
    words = text.split()
    words = [w for w in words if len(w) >= 3]

    return ' '.join(words)


def run_topic_modelling(index, num_topics=5, num_words=10):
    """
    Run LDA topic modelling on indexed documents.

    Args:
        index: Document index
        num_topics: Number of topics to extract
        num_words: Number of top words per topic

    Returns:
        topics_index: Dictionary with topics and document assignments
    """
    if not SKLEARN_AVAILABLE:
        return None, "scikit-learn not installed"

    documents = index.get("documents", {})

    if len(documents) < 2:
        return None, "Need at least 2 documents for topic modelling"

    # Prepare document texts
    doc_names = []
    doc_texts = []

    for doc_id, doc_data in documents.items():
        text = doc_data.get("text", "")
        if text and len(text) > 100:  # Only include documents with substantial text
            doc_names.append(doc_id)
            doc_texts.append(preprocess_text_for_topics(text))

    if len(doc_texts) < 2:
        return None, "Not enough documents with text content"

    # Adjust num_topics if we have fewer documents
    num_topics = min(num_topics, len(doc_texts))

    try:
        # Create document-term matrix using CountVectorizer
        vectorizer = CountVectorizer(
            max_df=0.95,  # Ignore terms that appear in more than 95% of documents
            min_df=1,     # Include terms that appear in at least 1 document
            max_features=1000,  # Limit to top 1000 words
            stop_words='english'  # Remove common English stop words
        )

        doc_term_matrix = vectorizer.fit_transform(doc_texts)

        # Run LDA
        lda = LatentDirichletAllocation(
            n_components=num_topics,
            random_state=42,
            max_iter=20,
            learning_method='online'
        )

        doc_topic_matrix = lda.fit_transform(doc_term_matrix)

        # Get feature names (words)
        feature_names = vectorizer.get_feature_names_out()

        # Extract topics with top words
        topics = []
        for topic_idx, topic in enumerate(lda.components_):
            top_word_indices = topic.argsort()[:-num_words-1:-1]
            top_words = [feature_names[i] for i in top_word_indices]
            top_weights = [float(topic[i]) for i in top_word_indices]

            topics.append({
                "id": topic_idx,
                "name": f"Topic {topic_idx + 1}",
                "words": top_words,
                "weights": top_weights,
                "label": ", ".join(top_words[:3])  # Use top 3 words as label
            })

        # Assign topics to documents
        document_topics = {}
        for doc_idx, doc_name in enumerate(doc_names):
            topic_distribution = doc_topic_matrix[doc_idx]
            dominant_topic = int(topic_distribution.argmax())
            topic_score = float(topic_distribution[dominant_topic])

            document_topics[doc_name] = {
                "dominant_topic": dominant_topic,
                "topic_label": topics[dominant_topic]["label"],
                "confidence": round(topic_score * 100, 1),
                "all_topics": [round(float(score) * 100, 1) for score in topic_distribution]
            }

        topics_index = {
            "topics": topics,
            "document_topics": document_topics,
            "num_topics": num_topics,
            "created_at": datetime.now().isoformat()
        }

        save_topics(topics_index)
        return topics_index, None

    except Exception as e:
        return None, str(e)


# =============================================================================
# SENTIMENT ANALYSIS FUNCTIONS
# =============================================================================

def load_sentiment():
    """Load sentiment index from file."""
    if SENTIMENT_FILE.exists():
        with open(SENTIMENT_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"documents": {}, "summary": {}}


def save_sentiment(sentiment_index):
    """Save sentiment index to file."""
    with open(SENTIMENT_FILE, 'w', encoding='utf-8') as f:
        json.dump(sentiment_index, f, ensure_ascii=False, indent=2)


def analyze_sentiment(text):
    """
    Analyze sentiment of text using TextBlob.
    Returns polarity (-1 to 1) and subjectivity (0 to 1).
    """
    if not TEXTBLOB_AVAILABLE:
        return None, None

    try:
        blob = TextBlob(text)
        return blob.sentiment.polarity, blob.sentiment.subjectivity
    except:
        return 0, 0


def get_sentiment_label(polarity):
    """Convert polarity score to human-readable label."""
    if polarity >= 0.3:
        return "Positive", "🟢"
    elif polarity >= 0.1:
        return "Slightly Positive", "🟡"
    elif polarity > -0.1:
        return "Neutral", "⚪"
    elif polarity > -0.3:
        return "Slightly Negative", "🟠"
    else:
        return "Negative", "🔴"


def get_risk_level(polarity, subjectivity):
    """
    Determine risk level based on sentiment.
    For auditing: negative + subjective = higher risk
    """
    if polarity < -0.3:
        return "High Risk", "🔴"
    elif polarity < -0.1:
        return "Medium Risk", "🟠"
    elif polarity < 0.1 and subjectivity > 0.5:
        return "Review Needed", "🟡"
    else:
        return "Low Risk", "🟢"


# Risk keywords for audit context
RISK_KEYWORDS = {
    "high_risk": [
        "fraud", "embezzlement", "misappropriation", "violation", "illegal",
        "unauthorized", "forgery", "falsification", "corruption", "bribery",
        "theft", "loss", "missing", "discrepancy", "irregularity", "breach",
        "non-compliance", "misconduct", "negligence", "failure"
    ],
    "medium_risk": [
        "delay", "overdue", "pending", "incomplete", "error", "mistake",
        "unclear", "unverified", "undocumented", "missing documentation",
        "late submission", "deviation", "variance", "exception", "issue",
        "concern", "problem", "weakness", "deficiency"
    ],
    "attention": [
        "urgent", "important", "critical", "priority", "immediate",
        "attention required", "action needed", "follow-up", "review",
        "verify", "confirm", "investigate", "clarify"
    ]
}


def detect_risk_keywords(text):
    """Detect risk-related keywords in text."""
    text_lower = text.lower()
    found_keywords = {
        "high_risk": [],
        "medium_risk": [],
        "attention": []
    }

    for category, keywords in RISK_KEYWORDS.items():
        for keyword in keywords:
            if keyword in text_lower:
                # Count occurrences
                count = text_lower.count(keyword)
                found_keywords[category].append({"keyword": keyword, "count": count})

    return found_keywords


def run_sentiment_analysis(index):
    """
    Run sentiment analysis on all indexed documents.
    """
    if not TEXTBLOB_AVAILABLE:
        return None, "TextBlob not installed"

    documents = index.get("documents", {})

    if not documents:
        return None, "No documents to analyze"

    sentiment_index = {
        "documents": {},
        "summary": {
            "total": 0,
            "positive": 0,
            "neutral": 0,
            "negative": 0,
            "high_risk": 0,
            "medium_risk": 0,
            "low_risk": 0
        }
    }

    for doc_id, doc_data in documents.items():
        text = doc_data.get("text", "")

        if not text or len(text) < 50:
            continue

        # Analyze sentiment (limit text length for performance)
        sample_text = text[:10000] if len(text) > 10000 else text
        polarity, subjectivity = analyze_sentiment(sample_text)

        if polarity is None:
            continue

        # Get labels
        sentiment_label, sentiment_icon = get_sentiment_label(polarity)
        risk_label, risk_icon = get_risk_level(polarity, subjectivity)

        # Detect risk keywords
        risk_keywords = detect_risk_keywords(text)

        # Adjust risk based on keywords
        if risk_keywords["high_risk"]:
            risk_label, risk_icon = "High Risk", "🔴"
        elif risk_keywords["medium_risk"] and risk_label == "Low Risk":
            risk_label, risk_icon = "Medium Risk", "🟠"

        sentiment_index["documents"][doc_id] = {
            "polarity": round(polarity, 3),
            "subjectivity": round(subjectivity, 3),
            "sentiment_label": sentiment_label,
            "sentiment_icon": sentiment_icon,
            "risk_label": risk_label,
            "risk_icon": risk_icon,
            "risk_keywords": risk_keywords
        }

        # Update summary
        sentiment_index["summary"]["total"] += 1

        if polarity >= 0.1:
            sentiment_index["summary"]["positive"] += 1
        elif polarity <= -0.1:
            sentiment_index["summary"]["negative"] += 1
        else:
            sentiment_index["summary"]["neutral"] += 1

        if risk_label == "High Risk":
            sentiment_index["summary"]["high_risk"] += 1
        elif risk_label == "Medium Risk":
            sentiment_index["summary"]["medium_risk"] += 1
        else:
            sentiment_index["summary"]["low_risk"] += 1

    sentiment_index["created_at"] = datetime.now().isoformat()
    save_sentiment(sentiment_index)

    return sentiment_index, None


# =============================================================================
# AI SUMMARIZATION FUNCTIONS
# =============================================================================

# Summarization model (lazy loaded)
_summarizer_model = None
_summarizer_tokenizer = None

def get_summarizer():
    """Get or initialize the summarization model and tokenizer."""
    global _summarizer_model, _summarizer_tokenizer
    if _summarizer_model is None and TRANSFORMERS_AVAILABLE:
        try:
            # Using t5-small (~250MB) instead of BART (~1.6GB) for faster downloads
            model_name = "t5-small"
            _summarizer_tokenizer = AutoTokenizer.from_pretrained(model_name)
            _summarizer_model = AutoModelForSeq2SeqLM.from_pretrained(model_name)
        except Exception as e:
            return None, None
    return _summarizer_model, _summarizer_tokenizer


def summarize_text(text, max_length=150, min_length=30):
    """Generate summary of text using T5 model."""
    if not TRANSFORMERS_AVAILABLE:
        return "[Summarization not available - install transformers]"

    model, tokenizer = get_summarizer()
    if model is None or tokenizer is None:
        return "[Model loading failed]"

    # Chunk long text (model has token limit ~512 tokens for t5-small)
    max_chunk = 512
    if len(text) > max_chunk:
        text = text[:max_chunk]

    # Skip if text is too short
    if len(text) < 100:
        return "[Text too short to summarize]"

    try:
        # T5 requires "summarize: " prefix
        input_text = "summarize: " + text
        inputs = tokenizer(input_text, return_tensors="pt", max_length=512, truncation=True)
        summary_ids = model.generate(
            inputs["input_ids"],
            max_length=max_length,
            min_length=min_length,
            num_beams=4,
            early_stopping=True
        )
        summary = tokenizer.decode(summary_ids[0], skip_special_tokens=True)
        return summary
    except Exception as e:
        return f"[Summarization error: {str(e)}]"


def load_summaries():
    """Load summaries index from file."""
    if SUMMARIES_FILE.exists():
        with open(SUMMARIES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"documents": {}}


def save_summaries(summaries):
    """Save summaries index to file."""
    with open(SUMMARIES_FILE, 'w', encoding='utf-8') as f:
        json.dump(summaries, f, ensure_ascii=False, indent=2)


# =============================================================================
# RAG (RETRIEVAL AUGMENTED GENERATION) FUNCTIONS - Using FAISS
# =============================================================================

# Embedding model and FAISS index (lazy loaded)
_embedding_model = None
_faiss_index = None
_faiss_chunks = []  # Store chunks for retrieval
_faiss_metadata = []  # Store metadata (doc_id, chunk_idx)

FAISS_INDEX_FILE = BASE_DIR / "faiss_index.bin"
FAISS_DATA_FILE = BASE_DIR / "faiss_data.json"


def get_embedding_model():
    """Get or initialize the sentence transformer model."""
    global _embedding_model
    if _embedding_model is None and SENTENCE_TRANSFORMERS_AVAILABLE:
        try:
            _embedding_model = SentenceTransformer('all-MiniLM-L6-v2')
        except Exception as e:
            return None
    return _embedding_model


def get_faiss_index():
    """Get or initialize the FAISS index."""
    global _faiss_index, _faiss_chunks, _faiss_metadata
    if _faiss_index is None and FAISS_AVAILABLE:
        # Try to load existing index
        if FAISS_INDEX_FILE.exists() and FAISS_DATA_FILE.exists():
            try:
                _faiss_index = faiss.read_index(str(FAISS_INDEX_FILE))
                with open(FAISS_DATA_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    _faiss_chunks = data.get('chunks', [])
                    _faiss_metadata = data.get('metadata', [])
            except Exception:
                _faiss_index = None
    return _faiss_index


def save_faiss_index():
    """Save FAISS index and data to disk."""
    global _faiss_index, _faiss_chunks, _faiss_metadata
    if _faiss_index is not None:
        try:
            faiss.write_index(_faiss_index, str(FAISS_INDEX_FILE))
            with open(FAISS_DATA_FILE, 'w', encoding='utf-8') as f:
                json.dump({
                    'chunks': _faiss_chunks,
                    'metadata': _faiss_metadata
                }, f, ensure_ascii=False)
        except Exception:
            pass


def get_chroma_collection():
    """Compatibility function - returns FAISS index status."""
    return get_faiss_index()


def chunk_text(text, chunk_size=500, overlap=50):
    """Split text into overlapping chunks for embedding."""
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start = end - overlap
        if start < 0:
            break
    return chunks


def index_document_embeddings(doc_id, text):
    """Create embeddings for document chunks and store in FAISS."""
    global _faiss_index, _faiss_chunks, _faiss_metadata

    if not SENTENCE_TRANSFORMERS_AVAILABLE or not FAISS_AVAILABLE:
        return False

    model = get_embedding_model()
    if not model:
        return False

    try:
        chunks = chunk_text(text)
        if not chunks:
            return False

        # Get embeddings
        embeddings = model.encode(chunks)
        embeddings = np.array(embeddings).astype('float32')

        # Initialize or update FAISS index
        if _faiss_index is None:
            dimension = embeddings.shape[1]
            _faiss_index = faiss.IndexFlatL2(dimension)
            _faiss_chunks = []
            _faiss_metadata = []

        # Add to index
        _faiss_index.add(embeddings)

        # Store chunks and metadata
        for i, chunk in enumerate(chunks):
            _faiss_chunks.append(chunk)
            _faiss_metadata.append({"doc_id": doc_id, "chunk_idx": i})

        return True
    except Exception as e:
        return False


def search_similar_chunks(query, n_results=5):
    """Find most relevant document chunks for a query using FAISS."""
    global _faiss_index, _faiss_chunks, _faiss_metadata

    if not SENTENCE_TRANSFORMERS_AVAILABLE or not FAISS_AVAILABLE:
        return None

    model = get_embedding_model()
    index = get_faiss_index()

    if not model or index is None:
        return None

    try:
        # Check if index has any documents
        if index.ntotal == 0:
            return None

        # Get query embedding
        query_embedding = model.encode([query])
        query_embedding = np.array(query_embedding).astype('float32')

        # Search
        k = min(n_results, index.ntotal)
        distances, indices = index.search(query_embedding, k)

        # Format results similar to ChromaDB format
        documents = []
        metadatas = []
        for idx in indices[0]:
            if idx < len(_faiss_chunks):
                documents.append(_faiss_chunks[idx])
                metadatas.append(_faiss_metadata[idx])

        return {
            'documents': [documents],
            'metadatas': [metadatas],
            'distances': [distances[0].tolist()]
        }
    except Exception as e:
        return None


def generate_answer(query, context):
    """Generate answer using context. Uses simple extractive approach."""
    if not context:
        return "No relevant information found in the documents."
    return f"Based on the documents:\n\n{context}"


def index_all_embeddings(index):
    """Index embeddings for all documents in the index."""
    global _faiss_index, _faiss_chunks, _faiss_metadata

    if not SENTENCE_TRANSFORMERS_AVAILABLE or not FAISS_AVAILABLE:
        return 0, "Required libraries not available"

    # Reset index for fresh indexing
    _faiss_index = None
    _faiss_chunks = []
    _faiss_metadata = []

    indexed_count = 0
    for doc_id, doc_data in index.get("documents", {}).items():
        text = doc_data.get("text", "")
        if text and len(text) > 100:
            if index_document_embeddings(doc_id, text):
                indexed_count += 1

    # Save index to disk
    save_faiss_index()

    return indexed_count, None


# =============================================================================
# ANOMALY DETECTION FUNCTIONS
# =============================================================================

def extract_amounts_from_text(text):
    """Extract monetary amounts from text."""
    amounts = []
    # Match various currency formats
    patterns = [
        r'\$\s?([\d,]+(?:\.\d{2})?)',           # $1,234.56
        r'Rs\.?\s?([\d,]+(?:\.\d{2})?)',        # Rs. 1,234.56
        r'PKR\s?([\d,]+(?:\.\d{2})?)',          # PKR 1,234.56
        r'USD\s?([\d,]+(?:\.\d{2})?)',          # USD 1,234.56
        r'([\d,]+(?:\.\d{2})?)\s?(?:million)',  # 1.5 million
        r'([\d,]+(?:\.\d{2})?)\s?(?:billion)',  # 1.5 billion
        r'([\d,]+(?:\.\d{2})?)\s?(?:thousand)', # 1.5 thousand
        r'([\d,]+(?:\.\d{2})?)\s?(?:lac|lakh)', # 1.5 lac/lakh
        r'([\d,]+(?:\.\d{2})?)\s?(?:crore)',    # 1.5 crore
    ]
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                # Remove commas and convert to float
                amount = float(match.replace(',', ''))
                if amount > 0:  # Only positive amounts
                    amounts.append(amount)
            except (ValueError, AttributeError):
                pass
    return amounts


def detect_anomalies(index):
    """Detect anomalies in document amounts using statistical methods."""
    if not SKLEARN_AVAILABLE or not NUMPY_AVAILABLE:
        return {"error": "scikit-learn or numpy not installed"}

    from sklearn.ensemble import IsolationForest

    doc_amounts = {}
    all_amounts = []
    amount_to_doc = []  # Track which document each amount belongs to

    for doc_id, doc_data in index.get("documents", {}).items():
        text = doc_data.get("text", "")
        amounts = extract_amounts_from_text(text)
        doc_amounts[doc_id] = amounts
        for amount in amounts:
            all_amounts.append(amount)
            amount_to_doc.append(doc_id)

    if len(all_amounts) < 10:
        return {"error": f"Not enough data for anomaly detection (found {len(all_amounts)} amounts, need at least 10)"}

    try:
        # Use Isolation Forest for anomaly detection
        amounts_array = np.array(all_amounts).reshape(-1, 1)
        clf = IsolationForest(contamination=0.1, random_state=42)
        predictions = clf.fit_predict(amounts_array)

        # Find anomalous amounts
        anomalies = []
        for idx, (amount, doc_id, pred) in enumerate(zip(all_amounts, amount_to_doc, predictions)):
            if pred == -1:  # -1 indicates anomaly
                anomalies.append({
                    "document": doc_id,
                    "amount": amount,
                    "formatted_amount": f"${amount:,.2f}" if amount < 1000000 else f"${amount/1000000:.2f}M",
                    "type": "statistical_outlier"
                })

        # Sort by amount (largest first)
        anomalies.sort(key=lambda x: x["amount"], reverse=True)

        return {
            "anomalies": anomalies,
            "total_amounts": len(all_amounts),
            "anomaly_count": len(anomalies),
            "documents_analyzed": len(doc_amounts)
        }

    except Exception as e:
        return {"error": f"Anomaly detection error: {str(e)}"}


# =============================================================================
# STREAMLIT USER INTERFACE
# =============================================================================

def main():
    """Main application."""

    st.set_page_config(
        page_title="Audit Document Search",
        page_icon="🔍",
        layout="wide"
    )

    st.title("🔍 Audit Document Search Engine")
    st.markdown("*Advanced search with Named Entity Recognition (NER)*")

    # Sidebar
    st.sidebar.header("⚙️ Settings")

    # Load indexes
    index = load_index()
    entities_index = load_entities()
    doc_count = len(index.get("documents", {}))

    # Sidebar metrics
    st.sidebar.metric("📄 Documents Indexed", doc_count)
    total_words = sum(d.get("word_count", 0) for d in index.get("documents", {}).values())
    st.sidebar.metric("📝 Total Words", f"{total_words:,}")

    # NER Status
    if SPACY_AVAILABLE:
        st.sidebar.success("🧠 spaCy NER: Available")
    else:
        st.sidebar.info("🧠 NER: Pattern-based (spaCy not installed)")

    # OCR Status
    if TESSERACT_AVAILABLE and PDF2IMAGE_AVAILABLE and POPPLER_PATH:
        st.sidebar.success("📷 OCR: Available (Tesseract + Poppler)")
    elif TESSERACT_AVAILABLE and PDF2IMAGE_AVAILABLE:
        st.sidebar.warning("📷 OCR: Poppler not found in C:\\")
    elif TESSERACT_AVAILABLE:
        st.sidebar.warning("📷 OCR: pdf2image not installed")
    elif PDF2IMAGE_AVAILABLE:
        st.sidebar.warning("📷 OCR: pytesseract not installed")
    else:
        st.sidebar.info("📷 OCR: Not available")

    # AI Features Status
    if TRANSFORMERS_AVAILABLE:
        st.sidebar.success("🤖 AI Summarization: Available")
    else:
        st.sidebar.info("🤖 AI Summarization: Not installed")

    if SENTENCE_TRANSFORMERS_AVAILABLE and FAISS_AVAILABLE:
        st.sidebar.success("💬 Q&A Chat: Available (FAISS)")
    else:
        st.sidebar.info("💬 Q&A Chat: Not installed")

    # Sidebar: Index documents
    st.sidebar.subheader("📁 Index Documents")
    st.sidebar.caption(f"Folder: `{DOCUMENTS_FOLDER}`")

    if st.sidebar.button("🔄 Re-index with NER"):
        with st.spinner("Indexing documents and extracting entities..."):
            index, entities_index = index_all_documents(use_spacy=SPACY_AVAILABLE)
            st.sidebar.success(f"✅ Indexed {len(index['documents'])} documents!")
            st.rerun()

    # Sidebar: Upload documents
    st.sidebar.subheader("📤 Upload Documents")
    uploaded_files = st.sidebar.file_uploader(
        "Upload files",
        type=['pdf', 'docx', 'xlsx', 'txt'],
        accept_multiple_files=True,
        label_visibility="collapsed"
    )

    if uploaded_files:
        for uploaded_file in uploaded_files:
            save_path = DOCUMENTS_FOLDER / uploaded_file.name
            with open(save_path, 'wb') as f:
                f.write(uploaded_file.getbuffer())
            st.sidebar.success(f"✅ Saved: {uploaded_file.name}")

        with st.spinner("Indexing new documents..."):
            index, entities_index = index_all_documents(use_spacy=SPACY_AVAILABLE)
            st.rerun()

    # Search help
    st.sidebar.subheader("❓ Search Help")
    with st.sidebar.expander("How to Search"):
        st.markdown("""
        **Simple Search:** `procurement`

        **Boolean Search:**
        - `audit AND financial`
        - `audit OR review`
        - `audit NOT draft`

        **Wildcard:** `procur*`

        **Entity Search:** Select entity type and value
        """)

    # Create tabs for different features
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "🔎 Search", "👤 Entities (NER)", "📊 Statistics",
        "📋 Table Extraction", "🏷️ Topics", "⚠️ Sentiment",
        "🤖 AI Assistant"
    ])

    # ========================
    # TAB 1: SEARCH
    # ========================
    with tab1:
        st.header("Search Documents")

        col1, col2, col3 = st.columns([3, 1, 1])

        with col1:
            search_query = st.text_input(
                "Enter your search query",
                placeholder="Type keywords, use AND/OR/NOT, or wildcards (*)",
                label_visibility="collapsed"
            )

        with col2:
            search_type = st.selectbox(
                "Search Type",
                ["Simple", "Boolean (AND/OR/NOT)", "Wildcard (*)", "Fuzzy (Partial)"],
                label_visibility="collapsed"
            )

        with col3:
            case_sensitive = st.checkbox("Case Sensitive", value=False)

        if search_query:
            with st.spinner("Searching..."):
                if search_type == "Simple":
                    results = simple_search(search_query, index, case_sensitive)
                elif search_type == "Boolean (AND/OR/NOT)":
                    results = boolean_search(search_query, index, case_sensitive)
                elif search_type == "Wildcard (*)":
                    results = wildcard_search(search_query, index, case_sensitive)
                else:
                    results = fuzzy_search(search_query, index)

            stats = get_search_statistics(results, index)

            stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
            with stat_col1:
                st.metric("Documents Found", stats["documents_found"])
            with stat_col2:
                st.metric("Total Matches", stats["total_matches"])
            with stat_col3:
                st.metric("Match Rate", f"{stats['match_percentage']:.1f}%")
            with stat_col4:
                if results and EXCEL_SUPPORT:
                    excel_file = export_results_to_excel(results, search_query)
                    if excel_file:
                        st.download_button(
                            label="📥 Export",
                            data=excel_file,
                            file_name=f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

            st.divider()

            if results:
                for i, result in enumerate(results, 1):
                    with st.container():
                        col_a, col_b = st.columns([4, 1])
                        with col_a:
                            st.markdown(f"### {i}. 📄 {result['document']}")
                        with col_b:
                            st.markdown(f"**{result['matches']} matches**")

                        if result.get("query_terms"):
                            terms_str = ", ".join(f"`{t}`" for t in result["query_terms"][:5])
                            st.markdown(f"**Matched:** {terms_str}")

                        snippet = result.get('snippet', '')
                        if snippet:
                            highlighted = highlight_text(snippet, result.get("query_terms", []))
                            st.markdown(
                                f'<div style="background-color: #f0f0f0; padding: 10px; border-radius: 5px;">'
                                f'...{highlighted}...</div>',
                                unsafe_allow_html=True
                            )

                        st.caption(f"📁 {result['path']}")
                        st.divider()
            else:
                st.info("🔍 No documents found matching your search.")

    # ========================
    # TAB 2: ENTITIES (NER)
    # ========================
    with tab2:
        st.header("👤 Named Entity Recognition")
        st.markdown("*Automatically extracted persons, organizations, locations, dates, and amounts*")

        if not entities_index.get("global"):
            st.warning("⚠️ No entities extracted yet. Click 'Re-index with NER' in the sidebar.")
        else:
            # Entity summary
            col1, col2, col3, col4, col5 = st.columns(5)

            global_entities = entities_index.get("global", {})

            with col1:
                st.metric("👤 Persons", len(global_entities.get("persons", [])))
            with col2:
                st.metric("🏢 Organizations", len(global_entities.get("organizations", [])))
            with col3:
                st.metric("📍 Locations", len(global_entities.get("locations", [])))
            with col4:
                st.metric("📅 Dates", len(global_entities.get("dates", [])))
            with col5:
                st.metric("💰 Money", len(global_entities.get("money", [])))

            # Export entities button
            if EXCEL_SUPPORT:
                excel_entities = export_entities_to_excel(entities_index)
                if excel_entities:
                    st.download_button(
                        label="📥 Export All Entities to Excel",
                        data=excel_entities,
                        file_name=f"entities_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            st.divider()

            # Entity tabs
            ent_tab1, ent_tab2, ent_tab3, ent_tab4, ent_tab5 = st.tabs(
                ["👤 Persons", "🏢 Organizations", "📍 Locations", "📅 Dates", "💰 Money"]
            )

            with ent_tab1:
                st.subheader("👤 Persons Mentioned")
                persons = global_entities.get("persons", [])
                if persons:
                    for entity, count in persons[:30]:
                        col_e, col_c = st.columns([4, 1])
                        with col_e:
                            if st.button(f"🔍 {entity}", key=f"person_{entity}"):
                                st.session_state["entity_search"] = ("persons", entity)
                        with col_c:
                            st.write(f"{count}x")
                else:
                    st.info("No persons detected")

            with ent_tab2:
                st.subheader("🏢 Organizations Mentioned")
                orgs = global_entities.get("organizations", [])
                if orgs:
                    for entity, count in orgs[:30]:
                        col_e, col_c = st.columns([4, 1])
                        with col_e:
                            if st.button(f"🔍 {entity}", key=f"org_{entity}"):
                                st.session_state["entity_search"] = ("organizations", entity)
                        with col_c:
                            st.write(f"{count}x")
                else:
                    st.info("No organizations detected")

            with ent_tab3:
                st.subheader("📍 Locations Mentioned")
                locations = global_entities.get("locations", [])
                if locations:
                    for entity, count in locations[:30]:
                        col_e, col_c = st.columns([4, 1])
                        with col_e:
                            if st.button(f"🔍 {entity}", key=f"loc_{entity}"):
                                st.session_state["entity_search"] = ("locations", entity)
                        with col_c:
                            st.write(f"{count}x")
                else:
                    st.info("No locations detected")

            with ent_tab4:
                st.subheader("📅 Dates Mentioned")
                dates = global_entities.get("dates", [])
                if dates:
                    for entity, count in dates[:30]:
                        st.write(f"• {entity} ({count}x)")
                else:
                    st.info("No dates detected")

            with ent_tab5:
                st.subheader("💰 Monetary Values")
                money = global_entities.get("money", [])
                if money:
                    for entity, count in money[:30]:
                        st.write(f"• {entity} ({count}x)")
                else:
                    st.info("No monetary values detected")

            # Handle entity search
            if "entity_search" in st.session_state:
                entity_type, entity_value = st.session_state["entity_search"]
                st.divider()
                st.subheader(f"🔍 Documents containing: {entity_value}")

                results = entity_search(entity_type, entity_value, index, entities_index)

                if results:
                    for i, result in enumerate(results, 1):
                        st.markdown(f"**{i}. {result['document']}** - {result['matches']} mentions")
                        st.caption(f"📁 {result['path']}")
                else:
                    st.info("No documents found")

                if st.button("Clear Search"):
                    del st.session_state["entity_search"]
                    st.rerun()

    # ========================
    # TAB 3: STATISTICS
    # ========================
    with tab3:
        st.header("📊 Document Statistics")

        if doc_count > 0:
            col1, col2, col3 = st.columns(3)

            with col1:
                st.metric("Total Documents", doc_count)
            with col2:
                st.metric("Total Words", f"{total_words:,}")
            with col3:
                avg_words = total_words // doc_count if doc_count > 0 else 0
                st.metric("Avg Words/Doc", f"{avg_words:,}")

            st.divider()

            # Document list
            st.subheader("📚 Indexed Documents")

            doc_data_list = []
            for doc_id, doc_data in index.get("documents", {}).items():
                doc_entities = entities_index.get("documents", {}).get(doc_id, {})
                doc_data_list.append({
                    "Name": doc_data['name'],
                    "Words": doc_data.get('word_count', 0),
                    "Persons": len(doc_entities.get('persons', [])),
                    "Orgs": len(doc_entities.get('organizations', [])),
                    "Locations": len(doc_entities.get('locations', [])),
                    "Indexed": doc_data.get('indexed_at', 'Unknown')[:10]
                })

            st.dataframe(doc_data_list, use_container_width=True)
        else:
            st.info("📭 No documents indexed yet.")

    # ========================
    # TAB 4: TABLE EXTRACTION
    # ========================
    with tab4:
        st.header("📋 PDF Table Extraction")
        st.markdown("*Extract tables from multiple PDF files and export to Excel*")

        if not PDFPLUMBER_AVAILABLE:
            st.error("❌ pdfplumber not installed. Run: `pip install pdfplumber`")
        else:
            st.success("✅ Table extraction is available")

            # Upload PDFs for table extraction
            st.subheader("📤 Upload PDF Files")
            pdf_files = st.file_uploader(
                "Select PDF files to extract tables from",
                type=['pdf'],
                accept_multiple_files=True,
                key="table_pdf_upload"
            )

            if pdf_files:
                st.info(f"📄 {len(pdf_files)} PDF file(s) selected")

                # Export mode selection
                col1, col2 = st.columns(2)
                with col1:
                    export_mode = st.radio(
                        "Export Mode",
                        ["separate_sheets", "single_sheet"],
                        format_func=lambda x: "Each PDF in separate sheet" if x == "separate_sheets" else "All tables in one sheet"
                    )

                with col2:
                    extract_button = st.button("🔍 Extract Tables", type="primary")

                if extract_button:
                    with st.spinner("Extracting tables from PDFs..."):
                        all_tables = {}
                        errors = []
                        total_tables = 0

                        # Progress bar
                        progress_bar = st.progress(0)
                        status_text = st.empty()

                        for i, pdf_file in enumerate(pdf_files):
                            status_text.text(f"Processing: {pdf_file.name}")
                            progress_bar.progress((i + 1) / len(pdf_files))

                            tables, error = extract_tables_from_pdf(pdf_file)

                            if error:
                                errors.append(f"{pdf_file.name}: {error}")
                            elif tables:
                                all_tables[pdf_file.name] = tables
                                total_tables += len(tables)

                        progress_bar.empty()
                        status_text.empty()

                    # Show results
                    if errors:
                        for err in errors:
                            st.warning(f"⚠️ {err}")

                    if all_tables:
                        st.success(f"✅ Extracted {total_tables} table(s) from {len(all_tables)} file(s)")

                        # Store in session state for preview
                        st.session_state["extracted_tables"] = all_tables

                        # Export button
                        excel_file = export_tables_to_excel(all_tables, mode=export_mode)
                        if excel_file:
                            st.download_button(
                                label="📥 Download Excel File",
                                data=excel_file,
                                file_name=f"extracted_tables_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary"
                            )
                    else:
                        st.warning("⚠️ No tables found in the uploaded PDF files.")

            # Show preview of extracted tables
            if "extracted_tables" in st.session_state and st.session_state["extracted_tables"]:
                st.divider()
                st.subheader("👁️ Preview Extracted Tables")

                all_tables = st.session_state["extracted_tables"]

                for file_name, tables in all_tables.items():
                    with st.expander(f"📄 {file_name} ({len(tables)} tables)"):
                        for table_info in tables:
                            st.markdown(f"**Table {table_info['table_num']}** (Page {table_info['page']}) - {table_info['rows']} rows, {table_info['cols']} columns")

                            # Convert to dataframe-like display
                            if table_info['data']:
                                # Use first row as header if it looks like a header
                                data = table_info['data']
                                if len(data) > 1:
                                    st.dataframe(data, use_container_width=True)
                                else:
                                    st.write(data)

                            st.divider()

            # Instructions
            with st.expander("ℹ️ How to Use"):
                st.markdown("""
                **Steps:**
                1. Click "Browse files" to select one or more PDF files
                2. Choose export mode:
                   - **Separate sheets**: Each PDF gets its own Excel sheet
                   - **Single sheet**: All tables combined in one sheet
                3. Click "Extract Tables"
                4. Preview the extracted tables
                5. Click "Download Excel File" to save

                **Tips:**
                - Works best with clear, structured tables
                - Complex layouts may not extract perfectly
                - Scanned PDFs may not work (use OCR first)
                """)

    # ========================
    # TAB 5: TOPIC MODELLING
    # ========================
    with tab5:
        st.header("🏷️ Topic Modelling (LDA)")
        st.markdown("*Automatically discover themes and topics in your documents*")

        if not SKLEARN_AVAILABLE:
            st.error("❌ scikit-learn not installed. Run: `pip install scikit-learn`")
        else:
            st.success("✅ Topic modelling is available")

            # Load existing topics
            topics_index = load_topics()

            # Settings
            st.subheader("⚙️ Settings")
            col1, col2, col3 = st.columns(3)

            with col1:
                num_topics = st.slider("Number of Topics", min_value=2, max_value=10, value=5)
            with col2:
                num_words = st.slider("Words per Topic", min_value=5, max_value=20, value=10)
            with col3:
                run_button = st.button("🔍 Run Topic Analysis", type="primary")

            if run_button:
                if doc_count < 2:
                    st.warning("⚠️ Need at least 2 documents for topic modelling. Please index some documents first.")
                else:
                    with st.spinner("Analyzing topics... This may take a moment..."):
                        topics_index, error = run_topic_modelling(index, num_topics, num_words)

                    if error:
                        st.error(f"❌ Error: {error}")
                    else:
                        st.success(f"✅ Found {len(topics_index['topics'])} topics!")
                        st.rerun()

            # Display topics
            if topics_index.get("topics"):
                st.divider()
                st.subheader("📚 Discovered Topics")

                # Show topics with their words
                for topic in topics_index["topics"]:
                    with st.expander(f"🏷️ {topic['name']}: {topic['label']}", expanded=True):
                        # Word cloud style display
                        words_display = " • ".join([f"**{word}**" for word in topic['words']])
                        st.markdown(words_display)

                        # Find documents in this topic
                        docs_in_topic = []
                        for doc_name, doc_info in topics_index.get("document_topics", {}).items():
                            if doc_info["dominant_topic"] == topic["id"]:
                                docs_in_topic.append({
                                    "name": doc_name,
                                    "confidence": doc_info["confidence"]
                                })

                        if docs_in_topic:
                            st.markdown(f"**Documents ({len(docs_in_topic)}):**")
                            for doc in sorted(docs_in_topic, key=lambda x: x["confidence"], reverse=True):
                                st.write(f"• {doc['name']} ({doc['confidence']}% match)")

                st.divider()

                # Document-Topic Matrix
                st.subheader("📄 Document Topics")

                doc_topic_data = []
                for doc_name, doc_info in topics_index.get("document_topics", {}).items():
                    doc_topic_data.append({
                        "Document": doc_name,
                        "Main Topic": f"Topic {doc_info['dominant_topic'] + 1}",
                        "Topic Label": doc_info["topic_label"],
                        "Confidence": f"{doc_info['confidence']}%"
                    })

                if doc_topic_data:
                    st.dataframe(doc_topic_data, use_container_width=True)

                # Filter by topic
                st.subheader("🔍 Filter Documents by Topic")

                topic_options = ["All Topics"] + [f"Topic {t['id']+1}: {t['label']}" for t in topics_index["topics"]]
                selected_topic = st.selectbox("Select Topic", topic_options)

                if selected_topic != "All Topics":
                    topic_id = int(selected_topic.split(":")[0].replace("Topic ", "")) - 1

                    filtered_docs = []
                    for doc_name, doc_info in topics_index.get("document_topics", {}).items():
                        if doc_info["dominant_topic"] == topic_id:
                            filtered_docs.append(doc_name)

                    if filtered_docs:
                        st.write(f"**{len(filtered_docs)} documents in this topic:**")
                        for doc in filtered_docs:
                            st.write(f"• {doc}")
                    else:
                        st.info("No documents found in this topic")

            else:
                st.info("👆 Click 'Run Topic Analysis' to discover topics in your documents")

            # Help section
            with st.expander("ℹ️ About Topic Modelling"):
                st.markdown("""
                **What is Topic Modelling?**

                Topic Modelling (LDA - Latent Dirichlet Allocation) automatically discovers hidden themes in your documents.

                **How it works:**
                1. Analyzes word patterns across all documents
                2. Groups frequently co-occurring words into "topics"
                3. Assigns each document to its most relevant topic

                **Example Topics:**
                - **Financial**: budget, expenditure, payment, audit, amount
                - **Procurement**: tender, contract, vendor, supply, purchase
                - **HR**: employee, salary, leave, staff, department

                **Tips:**
                - More documents = better topics
                - Adjust number of topics based on document variety
                - Topics are based on word patterns, not meaning
                """)

    # ========================
    # TAB 6: SENTIMENT ANALYSIS
    # ========================
    with tab6:
        st.header("⚠️ Sentiment & Risk Analysis")
        st.markdown("*Detect positive/negative content and flag high-risk documents*")

        if not TEXTBLOB_AVAILABLE:
            st.error("❌ TextBlob not installed. Run: `pip install textblob`")
        else:
            st.success("✅ Sentiment analysis is available")

            # Load existing sentiment data
            sentiment_index = load_sentiment()

            # Run analysis button
            col1, col2 = st.columns([1, 3])
            with col1:
                if st.button("🔍 Analyze Sentiment", type="primary"):
                    if doc_count == 0:
                        st.warning("⚠️ No documents indexed. Please index some documents first.")
                    else:
                        with st.spinner("Analyzing document sentiment..."):
                            sentiment_index, error = run_sentiment_analysis(index)

                        if error:
                            st.error(f"❌ Error: {error}")
                        else:
                            st.success(f"✅ Analyzed {sentiment_index['summary']['total']} documents!")
                            st.rerun()

            # Display results
            if sentiment_index.get("summary", {}).get("total", 0) > 0:
                st.divider()

                # Summary metrics
                st.subheader("📊 Summary")
                summary = sentiment_index["summary"]

                col1, col2, col3, col4, col5, col6 = st.columns(6)
                with col1:
                    st.metric("Total Analyzed", summary["total"])
                with col2:
                    st.metric("🟢 Positive", summary["positive"])
                with col3:
                    st.metric("⚪ Neutral", summary["neutral"])
                with col4:
                    st.metric("🔴 Negative", summary["negative"])
                with col5:
                    st.metric("⚠️ High Risk", summary["high_risk"])
                with col6:
                    st.metric("🟠 Medium Risk", summary["medium_risk"])

                st.divider()

                # Risk-based view
                st.subheader("🚨 Documents by Risk Level")

                risk_filter = st.selectbox(
                    "Filter by Risk Level",
                    ["All", "High Risk", "Medium Risk", "Low Risk", "Review Needed"]
                )

                # Prepare document list
                doc_sentiment_list = []
                for doc_id, sent_data in sentiment_index.get("documents", {}).items():
                    if risk_filter == "All" or sent_data["risk_label"] == risk_filter:
                        # Count risk keywords
                        high_risk_count = sum(k["count"] for k in sent_data["risk_keywords"]["high_risk"])
                        medium_risk_count = sum(k["count"] for k in sent_data["risk_keywords"]["medium_risk"])

                        doc_sentiment_list.append({
                            "Document": doc_id,
                            "Risk": f"{sent_data['risk_icon']} {sent_data['risk_label']}",
                            "Sentiment": f"{sent_data['sentiment_icon']} {sent_data['sentiment_label']}",
                            "Polarity": sent_data["polarity"],
                            "High Risk Keywords": high_risk_count,
                            "Medium Risk Keywords": medium_risk_count
                        })

                # Sort by risk (high risk first)
                risk_order = {"High Risk": 0, "Medium Risk": 1, "Review Needed": 2, "Low Risk": 3}
                doc_sentiment_list.sort(key=lambda x: risk_order.get(x["Risk"].split(" ", 1)[1], 4))

                if doc_sentiment_list:
                    st.dataframe(doc_sentiment_list, use_container_width=True)
                else:
                    st.info("No documents match the selected filter.")

                st.divider()

                # Detailed view
                st.subheader("📄 Document Details")

                # Select document to view details
                doc_options = list(sentiment_index.get("documents", {}).keys())
                if doc_options:
                    selected_doc = st.selectbox("Select Document", doc_options)

                    if selected_doc:
                        doc_sent = sentiment_index["documents"][selected_doc]

                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.markdown(f"**Risk Level:** {doc_sent['risk_icon']} {doc_sent['risk_label']}")
                        with col2:
                            st.markdown(f"**Sentiment:** {doc_sent['sentiment_icon']} {doc_sent['sentiment_label']}")
                        with col3:
                            st.markdown(f"**Polarity:** {doc_sent['polarity']:.2f}")

                        # Show risk keywords
                        st.markdown("**Risk Keywords Found:**")

                        if doc_sent["risk_keywords"]["high_risk"]:
                            high_risk_words = ", ".join([f"**{k['keyword']}** ({k['count']}x)"
                                                        for k in doc_sent["risk_keywords"]["high_risk"]])
                            st.markdown(f"🔴 High Risk: {high_risk_words}")

                        if doc_sent["risk_keywords"]["medium_risk"]:
                            medium_risk_words = ", ".join([f"**{k['keyword']}** ({k['count']}x)"
                                                          for k in doc_sent["risk_keywords"]["medium_risk"]])
                            st.markdown(f"🟠 Medium Risk: {medium_risk_words}")

                        if doc_sent["risk_keywords"]["attention"]:
                            attention_words = ", ".join([f"**{k['keyword']}** ({k['count']}x)"
                                                        for k in doc_sent["risk_keywords"]["attention"]])
                            st.markdown(f"🟡 Attention: {attention_words}")

                        if not any([doc_sent["risk_keywords"]["high_risk"],
                                   doc_sent["risk_keywords"]["medium_risk"],
                                   doc_sent["risk_keywords"]["attention"]]):
                            st.markdown("✅ No risk keywords detected")

            else:
                st.info("👆 Click 'Analyze Sentiment' to detect risk and sentiment in your documents")

            # Help section
            with st.expander("ℹ️ About Sentiment & Risk Analysis"):
                st.markdown("""
                **What does this do?**

                Analyzes documents for:
                1. **Sentiment** - Positive, Neutral, or Negative tone
                2. **Risk Level** - Based on sentiment and keywords

                **Risk Keywords Detected:**

                🔴 **High Risk:**
                fraud, embezzlement, violation, illegal, unauthorized, forgery, corruption, theft, missing, discrepancy

                🟠 **Medium Risk:**
                delay, overdue, pending, incomplete, error, unclear, unverified, deviation, weakness

                🟡 **Attention:**
                urgent, important, critical, priority, immediate, action needed, investigate

                **How to use:**
                1. Index your documents
                2. Click "Analyze Sentiment"
                3. Review high-risk documents first
                4. Check flagged keywords in each document
                """)

    # ========================
    # TAB 7: AI ASSISTANT
    # ========================
    with tab7:
        st.header("🤖 AI Assistant")
        st.markdown("*Intelligent audit assistance with AI-powered features*")

        ai_tab1, ai_tab2, ai_tab3 = st.tabs([
            "📝 Summarization", "💬 Q&A Chat", "🔍 Anomaly Detection"
        ])

        # ========================
        # AI SUB-TAB 1: SUMMARIZATION
        # ========================
        with ai_tab1:
            st.subheader("📝 Document Summarization")
            st.markdown("*Auto-generate summaries of audit documents using AI*")

            if not TRANSFORMERS_AVAILABLE:
                st.error("❌ transformers not installed. Run: `pip install transformers torch`")
                st.info("**Note:** First run will download the BART model (~1.6GB)")
            else:
                st.success("✅ Summarization available")

                # Load existing summaries
                summaries = load_summaries()

                # Document selector
                doc_options = list(index.get("documents", {}).keys())
                if doc_options:
                    selected_doc = st.selectbox(
                        "Select Document to Summarize",
                        doc_options,
                        key="sum_doc_select"
                    )

                    col1, col2 = st.columns([1, 3])
                    with col1:
                        if st.button("📝 Generate Summary", type="primary"):
                            with st.spinner("Generating summary... (first run downloads model)"):
                                text = index["documents"][selected_doc]["text"]
                                # Limit input text
                                summary = summarize_text(text[:5000])

                                # Save summary
                                summaries["documents"][selected_doc] = {
                                    "summary": summary,
                                    "generated_at": datetime.now().isoformat()
                                }
                                save_summaries(summaries)

                                st.success("✅ Summary generated!")
                                st.rerun()

                    # Show existing summary if available
                    if selected_doc in summaries.get("documents", {}):
                        st.divider()
                        st.markdown("### Summary")
                        summary_data = summaries["documents"][selected_doc]
                        st.info(summary_data["summary"])
                        st.caption(f"Generated: {summary_data.get('generated_at', 'Unknown')[:19]}")

                    # Show all summaries
                    if summaries.get("documents"):
                        st.divider()
                        st.subheader("📚 All Generated Summaries")

                        for doc_name, sum_data in summaries["documents"].items():
                            with st.expander(f"📄 {doc_name}"):
                                st.markdown(sum_data["summary"])
                                st.caption(f"Generated: {sum_data.get('generated_at', 'Unknown')[:19]}")
                else:
                    st.warning("⚠️ No documents indexed. Please index some documents first.")

        # ========================
        # AI SUB-TAB 2: Q&A CHAT
        # ========================
        with ai_tab2:
            st.subheader("💬 Ask Questions About Documents")
            st.markdown("*Query your documents using natural language (RAG)*")

            if not SENTENCE_TRANSFORMERS_AVAILABLE or not FAISS_AVAILABLE:
                st.error("❌ Required libraries not installed.")
                st.code("pip install sentence-transformers faiss-cpu", language="bash")
                st.info("**Note:** First run will download the embedding model (~90MB)")
            else:
                st.success("✅ Q&A Chat available (using FAISS)")

                # Check if embeddings are indexed
                faiss_index = get_faiss_index()
                if faiss_index is not None:
                    doc_count_in_db = faiss_index.ntotal
                    st.caption(f"📊 {doc_count_in_db} document chunks indexed in vector database")
                else:
                    doc_count_in_db = 0
                    st.caption("📊 No documents indexed yet")

                # Index embeddings button (always show)
                col1, col2 = st.columns([1, 3])
                with col1:
                    if st.button("🔄 Index Embeddings", type="secondary"):
                        with st.spinner("Indexing document embeddings..."):
                            indexed_count, error = index_all_embeddings(index)
                            if error:
                                st.error(f"Error: {error}")
                            else:
                                st.success(f"✅ Indexed {indexed_count} documents!")
                                st.rerun()

                st.divider()

                # Query input
                query = st.text_input(
                    "Ask a question about your documents:",
                    placeholder="e.g., What were the main audit findings?"
                )

                if query:
                    with st.spinner("Searching documents..."):
                        results = search_similar_chunks(query)

                        if results and results.get('documents') and results['documents'][0]:
                            st.markdown("### 📋 Relevant Information")

                            # Show top results
                            for i, (doc, metadata) in enumerate(zip(
                                results['documents'][0],
                                results['metadatas'][0]
                            )):
                                with st.expander(f"📄 Result {i+1} - {metadata.get('doc_id', 'Unknown')}", expanded=(i==0)):
                                    st.markdown(doc)

                            # Generate combined answer
                            st.divider()
                            st.markdown("### 💡 Combined Context")
                            context = "\n\n---\n\n".join(results['documents'][0][:3])
                            answer = generate_answer(query, context)
                            st.info(answer)

                        else:
                            st.warning("No relevant information found. Try indexing embeddings first.")

            # Help section
            with st.expander("ℹ️ How Q&A Works"):
                st.markdown("""
                **Retrieval-Augmented Generation (RAG):**

                1. Documents are split into chunks
                2. Each chunk is converted to a vector embedding
                3. Your question is also converted to a vector
                4. Most similar chunks are retrieved
                5. These chunks provide context for answering

                **Tips:**
                - Click "Index Embeddings" after adding new documents
                - Ask specific questions for better results
                - The system finds relevant excerpts, not exact answers
                """)

        # ========================
        # AI SUB-TAB 3: ANOMALY DETECTION
        # ========================
        with ai_tab3:
            st.subheader("🔍 Anomaly Detection")
            st.markdown("*Detect unusual amounts or patterns in documents*")

            if not SKLEARN_AVAILABLE:
                st.error("❌ scikit-learn not installed. Run: `pip install scikit-learn`")
            elif not NUMPY_AVAILABLE:
                st.error("❌ numpy not installed. Run: `pip install numpy`")
            else:
                st.success("✅ Anomaly detection available")

                st.markdown("""
                This feature:
                - Extracts monetary amounts from all documents
                - Uses **Isolation Forest** algorithm to detect statistical outliers
                - Flags unusually large or small amounts that may need review
                """)

                if st.button("🔍 Run Anomaly Detection", type="primary"):
                    with st.spinner("Analyzing documents for anomalies..."):
                        results = detect_anomalies(index)

                        if "error" in results:
                            st.warning(f"⚠️ {results['error']}")
                        else:
                            st.success(f"✅ Analyzed {results['total_amounts']} amounts from {results['documents_analyzed']} documents")

                            col1, col2 = st.columns(2)
                            with col1:
                                st.metric("Total Amounts Found", results['total_amounts'])
                            with col2:
                                st.metric("Anomalies Detected", results['anomaly_count'])

                            if results["anomalies"]:
                                st.divider()
                                st.subheader("⚠️ Detected Anomalies")

                                # Create dataframe for display
                                anomaly_data = []
                                for a in results["anomalies"]:
                                    anomaly_data.append({
                                        "Document": a["document"],
                                        "Amount": a["formatted_amount"],
                                        "Raw Value": a["amount"],
                                        "Type": a["type"]
                                    })

                                st.dataframe(anomaly_data, use_container_width=True)

                                st.markdown("**Note:** These amounts are flagged as statistical outliers compared to other amounts in your documents. Review them to determine if they represent actual issues.")
                            else:
                                st.info("✅ No anomalies detected. All amounts appear within normal ranges.")

            # Help section
            with st.expander("ℹ️ About Anomaly Detection"):
                st.markdown("""
                **How it works:**

                1. Extracts all monetary values from documents (various formats: $, Rs., PKR, etc.)
                2. Applies **Isolation Forest** algorithm
                3. Identifies values that are statistically unusual
                4. Flags the top 10% as potential anomalies

                **What it detects:**
                - Unusually large amounts
                - Unusually small amounts
                - Values that don't fit the normal pattern

                **Currency formats detected:**
                - USD: $1,234.56
                - PKR: Rs. 1,234.56 or PKR 1,234.56
                - Amounts with qualifiers: million, billion, lac, crore

                **Limitations:**
                - Requires at least 10 amounts for analysis
                - Statistical outliers may not be actual issues
                - Manual review is always recommended
                """)


if __name__ == "__main__":
    main()
